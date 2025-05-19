'''
This file is the Manager for the Parts Sandbox application.
It is responsible for managing the overall flow of the application, including GUI, handling user interactions, and acting as the "Owner" of the excel database.
'''
import openpyxl
import pandas as pd
import os
import logging
from QMFile import QMFile

logger = logging.getLogger(__name__)

class ApplicationManager:
    '''
    This class is the manager for the Parts Sandbox application.
    It is responsible for managing the overall flow of the application, including GUI, handling user interactions, and acting as the "Owner" of the excel database.
    '''
    def __init__(self):
        '''
        Initializes the application manager.
        '''
        self.start()

    def start(self):
        '''
        Starts the application manager.
        '''
        self.check_for_db()

    def check_for_db(self):
        '''
        Checks for the database file.
        If it does not exist, it creates a new one.
        '''
        try:
            logger.info("Checking for parts_sandbox.xlsx")
            self.workbook = openpyxl.load_workbook(r'excels/parts_sandbox.xlsx')
        except FileNotFoundError:
            logger.info("Database file not found. Creating a new one.")
            workbook = openpyxl.Workbook()
            workbook.save(r'excels/parts_sandbox.xlsx')

    def update_alias(self, file_path):
        '''
        Updates the alias in the database.

        This function processes the "Master Part List" sheet in the provided Excel file and updates the "aliases" sheet 
        in the `parts_sandbox.xlsx` database. It ensures that aliases from the "Master Part List" are added to the 
        "aliases" sheet only if they do not already exist.

        Parameters:
            file_path (str): The path to the Excel file containing the "Master Part List" sheet.

        Functionality:
            - Reads the "Master Part List" sheet from the provided Excel file.
            - Checks if each alias in the "Master Part List" already exists in the "aliases" sheet of the database.
            - If an alias does not exist, it adds the alias and its corresponding value to the "aliases" sheet.
            - The "aliases" sheet acts as a simple alias-to-value mapping.

        Assumptions:
            - The "Master Part List" sheet contains aliases in column A and their corresponding values in column B.
            - The "aliases" sheet in the `parts_sandbox.xlsx` file has the same structure: aliases in column A and values in column B.

        Raises:
            FileNotFoundError: If the provided file or the database file does not exist.
            KeyError: If the "Master Part List" sheet is missing in the provided file.
            Exception: For any other unexpected errors.

        Example:
            Given a "Master Part List" sheet with the following data:
                | Alias       | Value       |
                |-------------|-------------|
                | Part123     | Widget A    |
                | Part456     | Widget B    |

            And an "aliases" sheet with the following data:
                | Alias       | Value       |
                |-------------|-------------|
                | Part123     | Widget A    |

            After running this function, the "aliases" sheet will be updated to:
                | Alias       | Value       |
                |-------------|-------------|
                | Part123     | Widget A    |
                | Part456     | Widget B    |
        '''
        try:
            logger.info(f'Loading workbook at {file_path}')
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook['Master Part List']

            #Make a dataframe of the sheet
            df = pd.DataFrame(sheet.values)
            logger.debug(df.head())
        except FileNotFoundError:
            raise FileNotFoundError(f"File {file_path} not found.")

    def refresh_all_files(self):
        """
        Refreshes the database with all Quote Master files.
        Returns True if successful, False if there were any non-critical errors.
        Raises Exception for critical errors.
        """
        try:
            logger.info("Starting database refresh")
            excel_folder = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'excels'))
            sandbox_path = os.path.join(excel_folder, 'parts_sandbox.xlsx')
            
            # Get list of Quote Master files
            files = [
                os.path.join(excel_folder, f) 
                for f in os.listdir(excel_folder) 
                if f.endswith('.xlsx') and f != "parts_sandbox.xlsx"
            ]
            
            if not files:
                logger.warning("No Quote Master files found")
                return False
                
            # Load or create aliases sheet
            try:
                existing_aliases = pd.read_excel(sandbox_path, sheet_name='aliases')
                logger.debug("Loaded existing aliases sheet")
            except (FileNotFoundError, KeyError):
                existing_aliases = pd.DataFrame(columns=['alias', 'value'])
                logger.info("Created new aliases sheet")
            
            had_warnings = False
            all_aliases = []
            
            # Process each file
            for file in files:
                try:
                    logger.debug(f"Processing file: {file}")
                    qm_file = QMFile(file)
                    master_sheet = qm_file.load_master_sheet()
                    df = pd.DataFrame(master_sheet.values)
                    
                    # Skip header row and get alias/value columns
                    df.columns = df.iloc[0]
                    df = df.iloc[1:]
                    
                    if 'alias' in df.columns and 'value' in df.columns:
                        aliases_from_file = df[['alias', 'value']].copy()
                        all_aliases.append(aliases_from_file)
                        logger.debug(f"Successfully processed {len(aliases_from_file)} aliases from {file}")
                    else:
                        logger.warning(f"File {file} missing required columns")
                        had_warnings = True
                        
                except Exception as e:
                    logger.error(f"Error processing {file}: {str(e)}", exc_info=True)
                    had_warnings = True
                    continue
                finally:
                    if hasattr(qm_file, 'workbook'):
                        qm_file.workbook.close()
            
            if all_aliases:
                # Combine all new aliases
                new_aliases = pd.concat(all_aliases, ignore_index=True)
                new_aliases = new_aliases.drop_duplicates(subset=['alias'], keep='first')
                
                # Merge with existing aliases, keeping existing ones in case of conflict
                combined_aliases = pd.concat([existing_aliases, new_aliases], ignore_index=True)
                combined_aliases = combined_aliases.drop_duplicates(subset=['alias'], keep='first')
                
                # Save back to parts_sandbox.xlsx
                with pd.ExcelWriter(sandbox_path, engine='openpyxl', mode='a') as writer:
                    combined_aliases.to_excel(writer, sheet_name='aliases', index=False)
                
                logger.info(f"Successfully saved {len(combined_aliases)} aliases to database")
            
            return not had_warnings
            
        except Exception as e:
            logger.error("Critical error during refresh", exc_info=True)
            raise Exception(f"Error preparing master file: {str(e)}")
        finally:
            if hasattr(self, 'workbook'):
                self.workbook.close()